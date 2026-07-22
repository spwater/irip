#!/usr/bin/env python3
"""Deterministic particle-size fixture generator for the IRIP platform.

Generates 60 particle-size experiments across four batches, two instruments,
and two method versions, with deterministic data driven by a fixed seed.

Output includes Excel files (.xlsx), minimal PDF reports, ground-truth JSON,
and a manifest with SHA-256 digests.  Two byte-for-byte duplicate copies,
three self-check failures, and two moisture warnings are injected.

Usage:
    python examples/particle-size/generate.py --output <dir> --seed 20260715
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import random
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

# ── Constants ───────────────────────────────────────────────────────────

DEFAULT_SEED: int = 20260715
EXPERIMENT_COUNT: int = 60
DUPLICATE_COUNT: int = 2
SELF_CHECK_FAILURE_COUNT: int = 3
MOISTURE_WARNING_COUNT: int = 2
CURVE_POINTS: int = 32

BATCHES: list[str] = ["B01", "B02", "B03", "B04"]
INSTRUMENTS: list[str] = ["LaserSizer-3000", "SieveSet-A"]
METHODS: dict[str, str] = {
    "LaserSizer-3000": "ISO-13320-1:2009",
    "SieveSet-A": "ISO-2591-1:1988",
}
NAME_STYLES: list[str] = ["D50_um", "MedianDiameter_mm", "\u4e2d\u4f4d\u5f84"]
SOURCE_UNITS: list[str] = ["um", "mm"]

# 0-based indices for anomalies — deterministic, no overlap between groups
FAILURE_INDICES: list[int] = [10, 25, 40]
MOISTURE_WARNING_INDICES: list[int] = [5, 35]
DUPLICATE_SOURCE_INDICES: list[int] = [0, 30]

TIMESTAMP_BASE: datetime = datetime(2026, 1, 15, 8, 0, 0)
FIXED_ZIP_DT: tuple[int, int, int, int, int, int] = (2026, 1, 15, 8, 0, 0)


# ── Data Classes ─────────────────────────────────────────────────────────


@dataclass(frozen=True)
class ParticleFixtureSummary:
    """Summary of a generated particle-size fixture."""

    manifest_digest: str
    experiment_count: int
    duplicate_count: int
    self_check_failures: int
    moisture_warnings: int


@dataclass
class Experiment:
    """A single particle-size experiment with all metadata and distribution data."""

    index: int
    exp_id: str
    batch: str
    instrument: str
    method: str
    source_name_style: str
    source_unit: str
    operator: str
    sample_id: str
    timestamp: str
    d10: float
    d50: float
    d90: float
    specific_surface: float
    moisture: float
    size_bins: list[float]
    cumulative_passing: list[float]
    self_check_failure: bool
    moisture_warning: bool
    is_duplicate_of: str | None


# ── PDF Generation ──────────────────────────────────────────────────────


def _generate_minimal_pdf(lines: list[str]) -> bytes:
    """Generate a minimal valid PDF 1.4 document containing *lines* as text.

    The PDF uses a single Helvetica font on a single US-Letter page.  No
    external libraries are required — the entire structure is built by hand
    with deterministic byte offsets.
    """
    # Build the content stream (text-showing operators).
    content_parts: list[str] = ["BT", "/F1 12 Tf", "50 750 Td"]
    for idx, line in enumerate(lines):
        escaped = (
            line.replace("\\", "\\\\")
            .replace("(", "\\(")
            .replace(")", "\\)")
        )
        content_parts.append(f"({escaped}) Tj")
        if idx < len(lines) - 1:
            content_parts.append("0 -15 Td")
    content_parts.append("ET")
    content_stream = "\n".join(content_parts).encode("latin-1")

    # Five indirect objects: Catalog, Pages, Page, Content stream, Font.
    objects: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        (
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792]"
            b" /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>"
        ),
        (
            b"<< /Length "
            + str(len(content_stream)).encode("latin-1")
            + b" >>\nstream\n"
            + content_stream
            + b"\nendstream"
        ),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]

    # Assemble the PDF body and track byte offsets for the xref table.
    pdf = bytearray(b"%PDF-1.4\n")
    offsets: list[int] = []
    for i, obj in enumerate(objects):
        offsets.append(len(pdf))
        pdf.extend(f"{i + 1} 0 obj\n".encode("latin-1"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")

    # Cross-reference table.
    xref_offset = len(pdf)
    num_objs = len(objects) + 1  # +1 for the free entry (object 0)
    pdf.extend(f"xref\n0 {num_objs}\n".encode("latin-1"))
    pdf.extend(b"0000000000 65535 f \n")
    for off in offsets:
        pdf.extend(f"{off:010d} 00000 n \n".encode("latin-1"))

    # Trailer.
    pdf.extend(b"trailer\n")
    pdf.extend(f"<< /Size {num_objs} /Root 1 0 R >>\n".encode("latin-1"))
    pdf.extend(b"startxref\n")
    pdf.extend(f"{xref_offset}\n".encode("latin-1"))
    pdf.extend(b"%%EOF")
    return bytes(pdf)


def _build_pdf_for_experiment(exp: Experiment) -> bytes:
    """Build a minimal PDF report for a single experiment."""
    lines: list[str] = [
        f"{exp.exp_id} Particle Size Report",
        "=" * 40,
        f"Batch: {exp.batch}",
        f"Instrument: {exp.instrument}",
        f"Method: {exp.method}",
        f"Operator: {exp.operator}",
        f"Sample ID: {exp.sample_id}",
        f"Timestamp: {exp.timestamp}",
        f"D10: {exp.d10} {exp.source_unit}",
        f"D50: {exp.d50} {exp.source_unit}",
        f"D90: {exp.d90} {exp.source_unit}",
        f"Specific Surface: {exp.specific_surface} m2/kg",
        f"Moisture: {exp.moisture} %",
        f"Self Check: {'FAIL' if exp.self_check_failure else 'PASS'}",
        f"Moisture Warning: {'YES' if exp.moisture_warning else 'NO'}",
        f"Curve Points: {len(exp.cumulative_passing)}",
    ]
    return _generate_minimal_pdf(lines)


# ── Excel Generation ────────────────────────────────────────────────────


_FIXED_MODIFIED_RE = re.compile(
    r"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)"
)
_FIXED_MODIFIEDReplacement = r"\g<1>2026-01-15T08:00:00Z\g<2>"


def _save_xlsx_deterministic(wb: Workbook, filepath: Path) -> None:
    """Save an openpyxl workbook with deterministic ZIP timestamps.

    Two sources of non-determinism are eliminated:

    1. **ZIP entry timestamps** — ``zipfile.ZipInfo`` defaults to the current
       wall-clock time.  We re-zip with a fixed ``date_time`` for every entry.
    2. **docProps/core.xml modified field** — openpyxl's ``save_workbook``
       explicitly sets ``properties.modified = datetime.now(...)`` right
       before serialising.  We patch the XML content to a fixed value.
    """
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    with zipfile.ZipFile(buffer, "r") as src:
        with zipfile.ZipFile(str(filepath), "w", zipfile.ZIP_DEFLATED) as dst:
            for info in sorted(src.infolist(), key=lambda i: i.filename):
                data = src.read(info.filename)
                if info.filename == "docProps/core.xml":
                    text = data.decode("utf-8")
                    text = _FIXED_MODIFIED_RE.sub(
                        _FIXED_MODIFIEDReplacement, text
                    )
                    data = text.encode("utf-8")
                new_info = zipfile.ZipInfo(info.filename, date_time=FIXED_ZIP_DT)
                new_info.compress_type = info.compress_type
                new_info.external_attr = info.external_attr
                dst.writestr(new_info, data)


def _write_excel(filepath: Path, exp: Experiment) -> None:
    """Write a single experiment to an Excel file with two sheets.

    Sheet 1 — *Summary*: key-value metadata pairs.
    Sheet 2 — *Distribution*: 32 rows of size bin vs cumulative % passing.
    """
    wb = Workbook()

    # Deterministic document properties.
    wb.properties.creator = "IRIP Fixture Generator"
    wb.properties.lastModifiedBy = "IRIP Fixture Generator"
    wb.properties.created = TIMESTAMP_BASE
    wb.properties.modified = TIMESTAMP_BASE
    wb.properties.title = f"Particle Size Report {exp.exp_id}"

    # Sheet 1: Summary.
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    ws.append(["Experiment ID", exp.exp_id])
    ws.append(["Batch", exp.batch])
    ws.append(["Instrument", exp.instrument])
    ws.append(["Method", exp.method])
    ws.append(["Source Name Style", exp.source_name_style])
    ws.append(["Source Unit", exp.source_unit])
    ws.append(["Operator", exp.operator])
    ws.append(["Sample ID", exp.sample_id])
    ws.append(["Timestamp", exp.timestamp])
    ws.append(["D10", exp.d10])
    ws.append(["D50", exp.d50])
    ws.append(["D90", exp.d90])
    ws.append(["Specific Surface (m2/kg)", exp.specific_surface])
    ws.append(["Moisture (%)", exp.moisture])
    ws.append(["Self Check Failure", str(exp.self_check_failure)])
    ws.append(["Moisture Warning", str(exp.moisture_warning)])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Sheet 2: Distribution.
    ws2 = wb.create_sheet("Distribution")
    ws2.append(["Size Bin", "Cumulative % Passing"])
    for size, pct in zip(exp.size_bins, exp.cumulative_passing, strict=False):
        ws2.append([size, pct])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    _save_xlsx_deterministic(wb, filepath)


# ── Experiment Generation ────────────────────────────────────────────────


def _generate_experiments(rng: random.Random) -> list[Experiment]:
    """Generate 60 deterministic particle-size experiments.

    All random values are drawn from *rng* (a ``random.Random`` seeded by the
    caller).  The call order is identical for every experiment regardless of
    anomaly flags, so the RNG sequence is fully deterministic.
    """
    experiments: list[Experiment] = []

    for i in range(EXPERIMENT_COUNT):
        exp_id = f"EXP-{i + 1:03d}"
        batch = BATCHES[i % len(BATCHES)]
        instrument = INSTRUMENTS[i % len(INSTRUMENTS)]
        method = METHODS[instrument]
        source_name_style = NAME_STYLES[i % len(NAME_STYLES)]
        source_unit = SOURCE_UNITS[i % len(SOURCE_UNITS)]
        timestamp = (
            TIMESTAMP_BASE + timedelta(hours=i * 2, minutes=(i * 13) % 60)
        ).isoformat()

        # ── Always call rng in the same order (determinism) ──
        operator = f"OP-{rng.randint(1, 5):02d}"
        sample_id = f"S-{rng.randint(1000, 9999)}"
        d10 = round(rng.uniform(0.5, 5.0), 2)
        d50 = round(rng.uniform(5.0, 50.0), 2)
        d90 = round(rng.uniform(50.0, 200.0), 2)
        specific_surface = round(rng.uniform(0.1, 10.0), 2)
        moisture = round(rng.uniform(0.1, 5.0), 2)
        curve_raw = sorted(
            round(rng.uniform(0.0, 100.0), 2) for _ in range(CURVE_POINTS)
        )

        # 32-point logarithmically spaced size bins.
        min_size = max(d10 / 10.0, 0.01)
        max_size = d90 * 10.0
        ratio = max_size / min_size
        size_bins = [
            round(min_size * ratio ** (j / (CURVE_POINTS - 1)), 4)
            for j in range(CURVE_POINTS)
        ]

        # ── Anomaly flags (no additional rng calls) ──
        is_failure = i in FAILURE_INDICES
        is_warning = i in MOISTURE_WARNING_INDICES

        if is_failure:
            if i == FAILURE_INDICES[0]:
                # Swap d10 and d50 → d10 > d50 (violates D10 < D50)
                d10, d50 = d50, d10
            elif i == FAILURE_INDICES[1]:
                # Swap d50 and d90 → d50 > d90 (violates D50 < D90)
                d50, d90 = d90, d50
            elif i == FAILURE_INDICES[2]:
                # Swap d10 and d50 → d10 > d50 (violates D10 < D50)
                d10, d50 = d50, d10

        if is_warning:
            # Force moisture > 3.0 % without consuming rng state.
            moisture = round(3.0 + (i + 1) * 0.05, 2)

        experiments.append(
            Experiment(
                index=i,
                exp_id=exp_id,
                batch=batch,
                instrument=instrument,
                method=method,
                source_name_style=source_name_style,
                source_unit=source_unit,
                operator=operator,
                sample_id=sample_id,
                timestamp=timestamp,
                d10=d10,
                d50=d50,
                d90=d90,
                specific_surface=specific_surface,
                moisture=moisture,
                size_bins=size_bins,
                cumulative_passing=curve_raw,
                self_check_failure=is_failure,
                moisture_warning=is_warning,
                is_duplicate_of=None,
            )
        )

    return experiments


# ── Ground Truth & Manifest ─────────────────────────────────────────────


def _sha256_hex(data: bytes) -> str:
    """Return the hex SHA-256 digest of *data*."""
    return hashlib.sha256(data).hexdigest()


def _build_ground_truth(
    experiments: list[Experiment],
    duplicate_experiments: list[Experiment],
    output_dir: Path,
    seed: int,
) -> dict[str, Any]:
    """Build the ground-truth dictionary from the generated experiments.

    The ``experiments`` list contains 60 originals followed by 2 duplicates.
    Each entry includes the SHA-256 digest of its .xlsx file.
    """
    entries: list[dict[str, Any]] = []

    for exp in experiments + duplicate_experiments:
        xlsx_path = output_dir / f"{exp.exp_id}.xlsx"
        entries.append(
            {
                "id": exp.exp_id,
                "batch": exp.batch,
                "instrument": exp.instrument,
                "method": exp.method,
                "source_name_style": exp.source_name_style,
                "source_unit": exp.source_unit,
                "operator": exp.operator,
                "sample_id": exp.sample_id,
                "timestamp": exp.timestamp,
                "d10": exp.d10,
                "d50": exp.d50,
                "d90": exp.d90,
                "specific_surface": exp.specific_surface,
                "moisture": exp.moisture,
                "is_duplicate_of": exp.is_duplicate_of,
                "self_check_failure": exp.self_check_failure,
                "moisture_warning": exp.moisture_warning,
                "file_digest": "sha256:" + _sha256_hex(xlsx_path.read_bytes()),
            }
        )

    return {
        "seed": seed,
        "experiment_count": EXPERIMENT_COUNT,
        "duplicate_count": DUPLICATE_COUNT,
        "self_check_failures": SELF_CHECK_FAILURE_COUNT,
        "moisture_warnings": MOISTURE_WARNING_COUNT,
        "experiments": entries,
    }


def _build_manifest(
    experiments: list[Experiment],
    duplicate_experiments: list[Experiment],
    output_dir: Path,
) -> tuple[dict[str, Any], str]:
    """Build the manifest dictionary and compute its SHA-256 digest.

    The digest is computed over the manifest JSON **without** the
    ``manifest_digest`` key, using ``sort_keys=True`` and compact separators
    to guarantee a canonical byte representation.
    """
    files_info: list[dict[str, Any]] = []

    # Original xlsx + pdf files.
    for exp in experiments:
        for suffix, kind in ((".xlsx", "excel"), (".pdf", "pdf")):
            path = output_dir / f"{exp.exp_id}{suffix}"
            data = path.read_bytes()
            files_info.append(
                {
                    "path": path.name,
                    "sha256": _sha256_hex(data),
                    "size": len(data),
                    "kind": kind,
                }
            )

    # Duplicate xlsx files.
    for dup in duplicate_experiments:
        path = output_dir / f"{dup.exp_id}.xlsx"
        data = path.read_bytes()
        files_info.append(
            {
                "path": path.name,
                "sha256": _sha256_hex(data),
                "size": len(data),
                "kind": "excel",
            }
        )

    # Sort by path name for a deterministic file ordering.
    files_info.sort(key=lambda f: f["path"])

    manifest_base: dict[str, Any] = {
        "total_files": len(files_info),
        "experiments": EXPERIMENT_COUNT,
        "duplicates": DUPLICATE_COUNT,
        "files": files_info,
    }

    # Canonical JSON encoding for the digest.
    encoded = json.dumps(
        manifest_base, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    digest = "sha256:" + hashlib.sha256(encoded).hexdigest()

    manifest = {**manifest_base, "manifest_digest": digest}
    return manifest, digest


# ── Main Entry Point ────────────────────────────────────────────────────


def generate_particle_fixture(
    output: Path,
    seed: int = DEFAULT_SEED,
) -> ParticleFixtureSummary:
    """Generate deterministic particle-size fixture files.

    Args:
        output: Directory to write the generated files.
        seed: Random seed for deterministic generation.

    Returns:
        A :class:`ParticleFixtureSummary` with the manifest digest and counts.
    """
    output = Path(output)
    output.mkdir(parents=True, exist_ok=True)

    # All randomness flows from this single seeded instance — no global state.
    rng = random.Random(seed)
    experiments = _generate_experiments(rng)

    # Write xlsx and pdf files for every experiment.
    for exp in experiments:
        _write_excel(output / f"{exp.exp_id}.xlsx", exp)
        pdf_bytes = _build_pdf_for_experiment(exp)
        (output / f"{exp.exp_id}.pdf").write_bytes(pdf_bytes)

    # Create byte-for-byte duplicate copies.
    duplicate_experiments: list[Experiment] = []
    for _dup_idx, src_idx in enumerate(DUPLICATE_SOURCE_INDICES):
        src_exp = experiments[src_idx]
        dup_id = f"{src_exp.exp_id}-DUP1"
        dup_exp = Experiment(
            index=src_exp.index,
            exp_id=dup_id,
            batch=src_exp.batch,
            instrument=src_exp.instrument,
            method=src_exp.method,
            source_name_style=src_exp.source_name_style,
            source_unit=src_exp.source_unit,
            operator=src_exp.operator,
            sample_id=src_exp.sample_id,
            timestamp=src_exp.timestamp,
            d10=src_exp.d10,
            d50=src_exp.d50,
            d90=src_exp.d90,
            specific_surface=src_exp.specific_surface,
            moisture=src_exp.moisture,
            size_bins=list(src_exp.size_bins),
            cumulative_passing=list(src_exp.cumulative_passing),
            self_check_failure=src_exp.self_check_failure,
            moisture_warning=src_exp.moisture_warning,
            is_duplicate_of=src_exp.exp_id,
        )
        src_xlsx = output / f"{src_exp.exp_id}.xlsx"
        dup_xlsx = output / f"{dup_id}.xlsx"
        dup_xlsx.write_bytes(src_xlsx.read_bytes())
        duplicate_experiments.append(dup_exp)

    # Build manifest (must happen before writing manifest.json / ground_truth.json
    # so those metadata files are not included in the file list).
    manifest, manifest_digest = _build_manifest(
        experiments, duplicate_experiments, output
    )

    # Write manifest.json.
    (output / "manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    # Build and write ground_truth.json.
    ground_truth = _build_ground_truth(
        experiments, duplicate_experiments, output, seed
    )
    (output / "ground_truth.json").write_text(
        json.dumps(ground_truth, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    return ParticleFixtureSummary(
        manifest_digest=manifest_digest,
        experiment_count=EXPERIMENT_COUNT,
        duplicate_count=DUPLICATE_COUNT,
        self_check_failures=SELF_CHECK_FAILURE_COUNT,
        moisture_warnings=MOISTURE_WARNING_COUNT,
    )


def main() -> None:
    """CLI entry point for the fixture generator."""
    parser = argparse.ArgumentParser(
        description="Generate deterministic particle-size fixture data for IRIP."
    )
    parser.add_argument(
        "--output",
        type=Path,
        required=True,
        help="Output directory for generated files.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=DEFAULT_SEED,
        help=f"Random seed (default: {DEFAULT_SEED}).",
    )
    args = parser.parse_args()

    summary = generate_particle_fixture(args.output, seed=args.seed)

    print(
        f"Generated {summary.experiment_count} experiments, "
        f"{summary.experiment_count + summary.duplicate_count} source files, "
        f"{summary.duplicate_count} duplicate copies, "
        f"{summary.self_check_failures} blocked, "
        f"{summary.moisture_warnings} warnings."
    )
    print(f"Manifest digest: {summary.manifest_digest[:16]}...")


if __name__ == "__main__":
    main()
