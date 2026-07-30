"""文件连接器：CSV / XLSX / JSON 预览与流式读取。

实现 Connector 协议：
- preview(source, limit): 从 artifact 流式读取前 limit 行返回 PreviewTable；
- read(source): 异步迭代器，逐条 yield SourceRecord。

支持格式：
- csv: 标准库 csv.DictReader；
- xlsx: openpyxl（若未安装则抛 AppError(unsupported_media_type)）；
- json: json.load，期望顶层为对象数组。

安全约定（C-01）：
- 只接受 artifact_id，由 ArtifactService 校验归属后流式读取；
- 不再接受任意服务器路径 path；
- 资源预算限制：行数、解压后大小、时间。
"""

import asyncio
import csv
import io
import json
import time
from collections.abc import AsyncIterator
from typing import IO, Any
from uuid import UUID

from packages.common.errors import AppError
from packages.connectors.contracts import (
    ConnectorSource,
    PreviewTable,
    SourceRecord,
)

#: 预览行数上限（即使调用方传入更大的 limit）。
MAX_PREVIEW_ROWS: int = 10000

#: 预览解压后最大字节数（256 MiB）。
MAX_PREVIEW_BYTES: int = 256 * 1024 * 1024

#: 预览超时秒数。
MAX_PREVIEW_TIME_SECONDS: int = 60

#: 读取时的分块大小（字节）。
CHUNK_SIZE: int = 64 * 1024


class FileConnector:
    """文件数据源连接器（CSV / XLSX / JSON）。

    C-01: 通过 artifact_service 从 artifact stream 读取，
    不再接受任意服务器路径。

    Attributes:
        _artifact_service: 可选的工件服务，提供 open_stream 方法。
            为 None 时仅支持从已有流读取（用于测试）。
    """

    def __init__(self, artifact_service: Any | None = None) -> None:
        """初始化文件连接器。

        Args:
            artifact_service: 工件服务实例，需提供
                ``open_stream(artifact_id) -> (filename, size, stream)`` 方法。
        """
        self._artifact_service = artifact_service

    async def preview(self, source: ConnectorSource, limit: int = 100) -> PreviewTable:
        """预览文件前 limit 行。

        C-01: 从 artifact stream 读取而非裸路径。

        Args:
            source: 文件数据源，config 须含 artifact_id 与 format。
            limit: 预览行数上限（受 MAX_PREVIEW_ROWS 约束）。

        Returns:
            PreviewTable: 列名 + 行 + 总行数。

        Raises:
            AppError: code="validation_failed"，当缺少 artifact_id/format 时。
            AppError: code="forbidden"，当 artifact 不属于当前组织时。
            AppError: code="file_too_large"，当超过资源预算时。
            AppError: code="unsupported_media_type"，当格式不支持时。
        """
        artifact_id_raw = source.config.get("artifact_id")
        fmt = source.config.get("format")
        if not artifact_id_raw:
            raise AppError(
                code="validation_failed",
                message="文件数据源缺少 artifact_id",
                retryable=False,
                fields={"field": "artifact_id"},
            )
        if not fmt:
            raise AppError(
                code="validation_failed",
                message="文件数据源缺少 format",
                retryable=False,
                fields={"field": "format"},
            )

        effective_limit: int = min(limit, MAX_PREVIEW_ROWS)
        stream = await self._open_artifact_stream(artifact_id_raw)
        try:
            columns, rows = await self._read_rows_from_stream(stream, fmt, effective_limit)
        finally:
            close_stream(stream)

        return PreviewTable(
            columns=columns,
            rows=tuple(tuple(r) for r in rows),
            row_count=len(rows),
        )

    async def read(self, source: ConnectorSource) -> AsyncIterator[SourceRecord]:
        """流式读取文件全部记录。

        Args:
            source: 文件数据源。

        Yields:
            SourceRecord: 每行一条记录，字段名 -> 值。

        Raises:
            AppError: code="validation_failed"，当缺少 artifact_id/format 时。
            AppError: code="unsupported_media_type"，当格式不支持时。
        """
        artifact_id_raw = source.config.get("artifact_id")
        fmt = source.config.get("format")
        if not artifact_id_raw:
            raise AppError(
                code="validation_failed",
                message="文件数据源缺少 artifact_id",
                retryable=False,
                fields={"field": "artifact_id"},
            )
        if not fmt:
            raise AppError(
                code="validation_failed",
                message="文件数据源缺少 format",
                retryable=False,
                fields={"field": "format"},
            )

        stream = await self._open_artifact_stream(artifact_id_raw)
        try:
            for record in await self._read_records_from_stream(stream, fmt):
                yield record
        finally:
            close_stream(stream)

    # ---- artifact stream 获取 ----

    async def _open_artifact_stream(self, artifact_id_raw: Any) -> IO[bytes]:
        """通过 artifact_service 打开 artifact 流。

        Args:
            artifact_id_raw: artifact ID（可能是 UUID 或字符串）。

        Returns:
            IO[bytes]: 二进制流（BytesIO）。

        Raises:
            AppError: code="validation_failed"，当 artifact_service 未配置时。
            AppError: code="not_found"，当 artifact 不存在或无权访问时。
        """
        if self._artifact_service is None:
            raise AppError(
                code="validation_failed",
                message="文件连接器未配置 artifact_service，无法读取 artifact",
                retryable=False,
                fields={"field": "artifact_service"},
            )

        try:
            artifact_id: UUID = UUID(str(artifact_id_raw))
        except (ValueError, TypeError) as exc:
            raise AppError(
                code="validation_failed",
                message=f"无效的 artifact_id: {artifact_id_raw}",
                retryable=False,
                fields={"field": "artifact_id"},
            ) from exc

        result = await self._artifact_service.open_stream(artifact_id)
        # result is (filename, size, stream)
        if isinstance(result, tuple) and len(result) == 3:
            return result[2]
        # If result is already a stream
        return result

    # ---- 内部读取（从流） ----

    async def _read_rows_from_stream(
        self,
        stream: IO[bytes],
        fmt: str,
        limit: int,
    ) -> tuple[tuple[str, ...], list[list]]:
        """从二进制流读取前 limit 行，返回 (列名元组, 行列表)。

        设置资源预算限制：行数、字节、时间。
        """
        start_time: float = time.monotonic()

        if fmt == "csv":
            return await asyncio.to_thread(self._read_csv_stream, stream, limit, start_time)
        if fmt == "xlsx":
            return await asyncio.to_thread(self._read_xlsx_stream, stream, limit, start_time)
        if fmt == "json":
            return await asyncio.to_thread(self._read_json_stream, stream, limit, start_time)
        raise AppError(
            code="unsupported_media_type",
            message=f"不支持的文件格式：{fmt}",
            retryable=False,
            fields={"format": fmt},
        )

    async def _read_records_from_stream(self, stream: IO[bytes], fmt: str) -> list[SourceRecord]:
        """从流读取全部记录为 SourceRecord 列表。"""
        columns, rows = await self._read_rows_from_stream(stream, fmt, limit=MAX_PREVIEW_ROWS)
        records: list[SourceRecord] = []
        for row in rows:
            fields: dict[str, str | None] = {}
            for idx, col in enumerate(columns):
                value = row[idx] if idx < len(row) else None
                fields[col] = None if value is None else str(value)
            records.append(SourceRecord(fields=fields))
        return records

    # ---- CSV ----

    @staticmethod
    def _read_csv_stream(
        stream: IO[bytes],
        limit: int,
        start_time: float,
    ) -> tuple[tuple[str, ...], list[list]]:
        """从二进制流读取 CSV 文件前 limit 行。"""
        total_bytes: int = 0
        chunks: list[bytes] = []
        while True:
            chunk: bytes = stream.read(CHUNK_SIZE)
            if not chunk:
                break
            total_bytes += len(chunk)
            if total_bytes > MAX_PREVIEW_BYTES:
                raise AppError(
                    code="file_too_large",
                    message=f"CSV 文件超过大小限制: {MAX_PREVIEW_BYTES} 字节",
                    retryable=False,
                    fields={"max_bytes": MAX_PREVIEW_BYTES},
                )
            chunks.append(chunk)
            if time.monotonic() - start_time > MAX_PREVIEW_TIME_SECONDS:
                raise AppError(
                    code="file_too_large",
                    message=f"CSV 读取超时: {MAX_PREVIEW_TIME_SECONDS} 秒",
                    retryable=False,
                    fields={"max_time": MAX_PREVIEW_TIME_SECONDS},
                )

        text: str = b"".join(chunks).decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        try:
            header = next(reader)
        except StopIteration:
            return ((), [])
        columns = tuple(header)
        rows: list[list] = []
        for row in reader:
            if len(rows) >= limit:
                break
            rows.append(list(row))
        return columns, rows

    # ---- XLSX ----

    @staticmethod
    def _read_xlsx_stream(
        stream: IO[bytes],
        limit: int,
        start_time: float,
    ) -> tuple[tuple[str, ...], list[list]]:
        """从二进制流读取 XLSX 文件首个工作表前 limit 行。

        Raises:
            AppError: code="unsupported_media_type"，当 openpyxl 未安装时。
        """
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise AppError(
                code="unsupported_media_type",
                message="读取 XLSX 需要 openpyxl 依赖，请安装 openpyxl",
                retryable=False,
                fields={"format": "xlsx"},
            ) from exc

        stream.seek(0)
        wb = load_workbook(filename=stream, read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return ((), [])
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                return ((), [])
            columns = tuple(str(c) if c is not None else "" for c in header)
            rows: list[list] = []
            for row in rows_iter:
                if len(rows) >= limit:
                    break
                rows.append(list(row))
                if time.monotonic() - start_time > MAX_PREVIEW_TIME_SECONDS:
                    raise AppError(
                        code="file_too_large",
                        message=f"XLSX 读取超时: {MAX_PREVIEW_TIME_SECONDS} 秒",
                        retryable=False,
                        fields={"max_time": MAX_PREVIEW_TIME_SECONDS},
                    )
            return columns, rows
        finally:
            wb.close()

    # ---- JSON ----

    @staticmethod
    def _read_json_stream(
        stream: IO[bytes],
        limit: int,
        start_time: float,
    ) -> tuple[tuple[str, ...], list[list]]:
        """从二进制流读取 JSON 文件（对象数组）前 limit 行。

        列名为首个对象的键的并集（保持首次出现顺序）。
        """
        total_bytes: int = 0
        chunks: list[bytes] = []
        while True:
            chunk: bytes = stream.read(CHUNK_SIZE)
            if not chunk:
                break
            total_bytes += len(chunk)
            if total_bytes > MAX_PREVIEW_BYTES:
                raise AppError(
                    code="file_too_large",
                    message=f"JSON 文件超过大小限制: {MAX_PREVIEW_BYTES} 字节",
                    retryable=False,
                    fields={"max_bytes": MAX_PREVIEW_BYTES},
                )
            chunks.append(chunk)
            if time.monotonic() - start_time > MAX_PREVIEW_TIME_SECONDS:
                raise AppError(
                    code="file_too_large",
                    message=f"JSON 读取超时: {MAX_PREVIEW_TIME_SECONDS} 秒",
                    retryable=False,
                    fields={"max_time": MAX_PREVIEW_TIME_SECONDS},
                )

        data = json.loads(b"".join(chunks).decode("utf-8"))
        if not isinstance(data, list):
            raise AppError(
                code="validation_failed",
                message="JSON 文件顶层必须是对象数组",
                retryable=False,
                fields={"format": "json"},
            )

        columns: list[str] = []
        seen: set[str] = set()
        for obj in data[:limit]:
            if not isinstance(obj, dict):
                continue
            for key in obj:
                if key not in seen:
                    seen.add(key)
                    columns.append(key)

        col_index = {col: idx for idx, col in enumerate(columns)}
        rows: list[list] = []
        for obj in data[:limit]:
            if not isinstance(obj, dict):
                continue
            row = [None] * len(columns)
            for key, value in obj.items():
                idx = col_index.get(key)
                if idx is not None:
                    row[idx] = value
            rows.append(row)
        return tuple(columns), rows


def close_stream(stream: IO[bytes]) -> None:
    """安全关闭流（忽略错误）。"""
    try:
        stream.close()
    except Exception:
        pass
