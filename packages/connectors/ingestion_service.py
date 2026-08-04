"""端到端摄入管线（IRIP Task 16，标准层空表清理后简化版）。

IngestionPipeline 将外部源文件转化为 L2 事实。标准层空表
（variable / variable_version / mapping_profile / mapping_profile_version /
fact_template / fact_template_version）已 DROP（migration 0057），
原映射/标准化/模板校验逻辑全部删除。管线阶段简化为：

1. download: 读取文件内容并计算 SHA-256 摘要（用于幂等去重）；
2. parse: 解析文件为字段-值字典（CSV/XLSX/JSON）；
3. dedup: 通过 SHA-256 作为 idempotency_key 检查重复；
4. persist_fact: 创建 Fact；
5. quality: 返回空通过结果（无观察值可评估）；
6. finalize: 返回摄入结果。
"""

from __future__ import annotations

import asyncio
import hashlib
import math
from dataclasses import dataclass
from pathlib import Path
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from packages.common.database import ScopedSessionMixin
from packages.common.errors import AppError
from packages.facts.quality import QualityAssessment, QualityEngine
from packages.facts.service import CreateFactCommand, FactService


@dataclass(frozen=True)
class IngestionResult:
    """摄入作业结果。

    Attributes:
        fact_ids: 创建/匹配的事实 ID 元组。
        deduplicated: 是否有重复文件去重。
        quality: 质量评估结果。
        blocked: 是否被质量规则阻断。
        warnings: 警告数。
        error: 失败原因（成功时 None）。
    """

    fact_ids: tuple[UUID, ...]
    deduplicated: bool
    quality: QualityAssessment
    blocked: bool
    warnings: int
    error: str | None


def _compute_sha256(file_path: Path) -> str:
    """计算文件 SHA-256 摘要。

    Args:
        file_path: 文件路径。

    Returns:
        str: 十六进制 SHA-256 摘要。
    """
    sha = hashlib.sha256()
    with open(file_path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            sha.update(chunk)
    return sha.hexdigest()


def _parse_xlsx_summary(file_path: Path) -> dict[str, str]:
    """解析 XLSX 文件 Summary 工作表为字段-值字典。

    Summary 工作表格式为 [Field, Value] 键值对行。

    Args:
        file_path: XLSX 文件路径。

    Returns:
        dict[str, str]: 字段名→值字符串映射。

    Raises:
        AppError: code="unsupported_media_type"，当 openpyxl 未安装时。
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise AppError(
            code="unsupported_media_type",
            message="读取 XLSX 需要 openpyxl 依赖",
            retryable=False,
            fields={"format": "xlsx"},
        ) from exc

    wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return {}
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}

        # 第一行是表头 ["Field", "Value"]
        fields: dict[str, str] = {}
        for row in rows[1:]:
            if row is None or len(row) < 2:
                continue
            key = row[0]
            value = row[1]
            if key is None:
                continue
            key_str = str(key).strip()
            if value is None:
                fields[key_str] = ""
            elif isinstance(value, float):
                # 避免浮点数精度问题：整数浮点转整数字符串
                if math.isclose(value, round(value)):
                    fields[key_str] = str(round(value))
                else:
                    fields[key_str] = str(value)
            else:
                fields[key_str] = str(value).strip()
        return fields
    finally:
        wb.close()


def _parse_csv(file_path: Path) -> dict[str, str]:
    """解析 CSV 文件为字段-值字典。

    假设 CSV 有两列（Field, Value）。

    Args:
        file_path: CSV 文件路径。

    Returns:
        dict[str, str]: 字段名→值字符串映射。
    """
    import csv

    fields: dict[str, str] = {}
    with open(file_path, newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        for i, row in enumerate(reader):
            if i == 0:
                continue  # 跳过表头
            if len(row) < 2:
                continue
            fields[row[0].strip()] = row[1].strip()
    return fields


def _parse_json(file_path: Path) -> dict[str, str]:
    """解析 JSON 文件为字段-值字典。

    假设 JSON 是一个键值对对象。

    Args:
        file_path: JSON 文件路径。

    Returns:
        dict[str, str]: 字段名→值字符串映射。
    """
    import json

    with open(file_path, encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, dict):
        raise AppError(
            code="validation_failed",
            message="JSON 文件必须是键值对对象",
            retryable=False,
            fields={"format": "json"},
        )
    return {str(k): str(v) if v is not None else "" for k, v in data.items()}


def _parse_file(file_path: Path) -> dict[str, str]:
    """根据文件扩展名解析为字段-值字典。

    Args:
        file_path: 文件路径。

    Returns:
        dict[str, str]: 字段名→值字符串映射。

    Raises:
        AppError: code="unsupported_media_type"，当格式不支持时。
    """
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx_summary(file_path)
    if suffix == ".csv":
        return _parse_csv(file_path)
    if suffix == ".json":
        return _parse_json(file_path)
    raise AppError(
        code="unsupported_media_type",
        message=f"不支持的文件格式：{suffix}",
        retryable=False,
        fields={"suffix": suffix},
    )


def _empty_passed_assessment() -> QualityAssessment:
    """构建空通过的质量评估结果。

    标准层删除后无观察值可供评估，摄入管线统一返回 passed。

    Returns:
        QualityAssessment: 空通过结果。
    """
    return QualityAssessment(
        results=(),
        overall_status="passed",
        summary={"passed": 0, "warning": 0, "blocked": 0},
    )


class IngestionPipeline(ScopedSessionMixin):
    """端到端摄入管线：下载→解析→去重→持久化→质量→完成。

    标准层空表清理后，映射/标准化逻辑已删除，管线仅做文件摘要去重
    与事实创建。质量评估返回空通过结果。

    Attributes:
        _factory: 异步会话工厂。
        _fact_service: 事实服务。
        _quality_engine: 质量评估引擎。
        _dept_id: 当前部门 ID。
        _actor_id: 当前操作人 ID。
    """

    def __init__(
        self,
        session_factory: async_sessionmaker[AsyncSession],
        fact_service: FactService,
        quality_engine: QualityEngine,
        department_id: UUID,
        actor_id: UUID | None = None,
    ) -> None:
        """初始化摄入管线。

        Args:
            session_factory: 异步会话工厂。
            fact_service: 事实服务（用于创建事实）。
            quality_engine: 质量评估引擎。
            department_id: 当前部门 ID。
            actor_id: 当前操作人 ID（可选）。
        """
        self._factory = session_factory
        self._fact_service = fact_service
        self._quality_engine = quality_engine
        self._dept_id = department_id
        self._actor_id = actor_id

    async def ingest_file(
        self,
        file_path: Path,
        object_id: UUID,
    ) -> IngestionResult:
        """摄入单个文件。

        流程：
        1. download: 读取文件内容，计算 SHA-256 摘要；
        2. parse: 解析为字段-值字典；
        3. dedup: 检查幂等键是否已存在；
        4. persist_fact: 创建 Fact；
        5. quality: 返回空通过结果；
        6. finalize: 返回摄入结果。

        去重：通过 SHA-256 作为 idempotency_key，重复文件返回已有事实。

        Args:
            file_path: 文件路径。
            object_id: 工业对象 ID。

        Returns:
            IngestionResult: 摄入结果。
        """
        try:
            # 1. download: 计算 SHA-256（同步 I/O，用 to_thread 避免阻塞）
            file_sha256 = await asyncio.to_thread(_compute_sha256, file_path)
            idempotency_key = f"sha256:{file_sha256}"

            # 3. dedup: 去重检查
            from packages.facts.repository import FactRepository

            async with self._scoped_session() as session:
                existing_fact = await FactRepository.find_by_idempotency_key(
                    session, self._dept_id, idempotency_key
                )

            if existing_fact is not None:
                # 重复文件：返回已有事实，标记去重
                return IngestionResult(
                    fact_ids=(existing_fact.id,),
                    deduplicated=True,
                    quality=_empty_passed_assessment(),
                    blocked=False,
                    warnings=0,
                    error=None,
                )

            # 2. parse: 解析文件（同步 I/O，用 to_thread 避免阻塞）
            parsed = await asyncio.to_thread(_parse_file, file_path)

            # 提取实验 ID 作为 subject_id
            subject_id = (
                parsed.get("Experiment ID") or parsed.get("experiment_id") or file_path.stem
            )

            # 4. persist_fact: 创建事实
            command = CreateFactCommand(
                fact_type="experiment_run",
                department_id=self._dept_id,
                object_id=object_id,
                subject_id=subject_id,
                started_at=None,
                ended_at=None,
                idempotency_key=idempotency_key,
                created_by=self._actor_id or self._dept_id,
            )

            ref = await self._fact_service.create(command)

            # 5. quality: 返回空通过结果（无观察值可评估）
            assessment = _empty_passed_assessment()

            return IngestionResult(
                fact_ids=(ref.fact_id,),
                deduplicated=False,
                quality=assessment,
                blocked=False,
                warnings=0,
                error=None,
            )
        except AppError as exc:
            return IngestionResult(
                fact_ids=(),
                deduplicated=False,
                quality=QualityAssessment(
                    results=(),
                    overall_status="blocked",
                    summary={"passed": 0, "warning": 0, "blocked": 0},
                ),
                blocked=True,
                warnings=0,
                error=f"{exc.code}: {exc.message}",
            )
        except Exception as exc:
            return IngestionResult(
                fact_ids=(),
                deduplicated=False,
                quality=QualityAssessment(
                    results=(),
                    overall_status="blocked",
                    summary={"passed": 0, "warning": 0, "blocked": 0},
                ),
                blocked=True,
                warnings=0,
                error=str(exc),
            )

    async def ingest_batch(
        self,
        file_paths: tuple[Path, ...],
        object_id: UUID,
    ) -> tuple[IngestionResult, ...]:
        """批量摄入多个文件。

        逐个文件调用 ingest_file，收集结果。个别文件失败不影响其他文件。

        Args:
            file_paths: 文件路径元组。
            object_id: 工业对象 ID。

        Returns:
            tuple[IngestionResult, ...]: 每个文件的摄入结果元组。
        """
        results: list[IngestionResult] = []
        for file_path in file_paths:
            result = await self.ingest_file(
                file_path=file_path,
                object_id=object_id,
            )
            results.append(result)
        return tuple(results)
