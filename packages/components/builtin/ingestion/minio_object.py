"""MinIO 对象读取组件。

通过 context.artifact_service 读取 S3/MinIO 中的对象，
按 media_type 解析为 ObservationTable。

参数：
- object_key: S3 对象 key（必填）。
- format: 数据格式（可选，auto/csv/json/excel，默认 auto 按推断）。
- delimiter: CSV 分隔符（可选）。
- encoding: 文件编码（可选，默认 utf-8）。
"""

import asyncio
import io
from typing import Any

from packages.common.errors import AppError
from packages.components.builtin.ingestion.csv_reader import _coerce
from packages.components.builtin.types import ObservationTable
from packages.components.sdk import ComponentContext, ComponentResult


class MinioObject:
    """MinIO/S3 对象读取组件。"""

    async def execute(
        self,
        context: ComponentContext,
        params: dict[str, Any],
    ) -> ComponentResult:
        """从 MinIO 读取对象并输出 ObservationTable。"""
        object_key: str = params["object_key"]
        fmt: str = params.get("format", "auto")
        delimiter: str | None = params.get("delimiter")
        encoding: str = params.get("encoding", "utf-8")

        artifact_service = context.artifact_service
        if artifact_service is None:
            raise AppError(
                code="missing_dependency",
                message="artifact_service 未注入",
                retryable=False,
                fields={},
            )

        # 通过 artifact_service 的 s3_repo 下载对象内容
        s3_repo = getattr(artifact_service, "_s3", None)
        if s3_repo is None:
            raise AppError(
                code="missing_dependency",
                message="artifact_service 缺少 s3_repo",
                retryable=False,
                fields={},
            )

        data: bytes = await asyncio.to_thread(s3_repo.get_object, object_key)

        # 推断格式
        if fmt == "auto":
            if object_key.endswith(".json"):
                fmt = "json"
            elif object_key.endswith(".csv") or object_key.endswith(".tsv"):
                fmt = "csv"
            elif object_key.endswith(".xlsx"):
                fmt = "excel"
            else:
                fmt = "csv"

        if fmt == "json":
            table = self._parse_json(data, object_key, encoding)
        elif fmt == "csv":
            table = self._parse_csv(data, object_key, encoding, delimiter)
        elif fmt == "excel":
            table = await self._parse_excel(data, object_key)
        else:
            raise AppError(
                code="unsupported_format",
                message=f"不支持的数据格式: {fmt}",
                retryable=False,
                fields={"format": fmt},
            )

        return ComponentResult(
            outputs={"observations": table},
            summary=f"从 MinIO 读取 {object_key}: {table.row_count()} 行",
            metadata={
                "row_count": table.row_count(),
                "column_count": table.column_count(),
                "object_key": object_key,
            },
        )

    def _parse_json(self, data: bytes, object_key: str, encoding: str) -> ObservationTable:
        """解析 JSON 数据。"""
        import json

        parsed: Any = json.loads(data.decode(encoding))
        records: list[dict[str, Any]]
        if isinstance(parsed, list):
            records = [r if isinstance(r, dict) else {"value": r} for r in parsed]
        elif isinstance(parsed, dict):
            records = [parsed]
        else:
            records = [{"value": parsed}]

        col_set: list[str] = []
        for rec in records:
            for k in rec:
                if k not in col_set:
                    col_set.append(k)
        columns: tuple[str, ...] = tuple(col_set)
        return ObservationTable(
            columns=columns,
            rows=tuple(records),
            source_locations=({"object_key": object_key},),
        )

    def _parse_csv(
        self,
        data: bytes,
        object_key: str,
        encoding: str,
        delimiter: str | None,
    ) -> ObservationTable:
        """解析 CSV 数据。"""
        import csv

        text = data.decode(encoding)
        reader = csv.reader(
            io.StringIO(text),
            delimiter=delimiter or ",",
        )
        all_rows = list(reader)
        if not all_rows:
            return ObservationTable()

        columns: tuple[str, ...] = tuple(all_rows[0])
        data_rows: list[dict[str, Any]] = []
        for _idx, row in enumerate(all_rows[1:], start=1):
            if not row:
                continue
            record: dict[str, Any] = {}
            for col_name, cell in zip(columns, row, strict=False):
                record[col_name] = _coerce(cell)
            data_rows.append(record)

        return ObservationTable(
            columns=columns,
            rows=tuple(data_rows),
            source_locations=({"object_key": object_key},),
        )

    async def _parse_excel(self, data: bytes, object_key: str) -> ObservationTable:
        """解析 Excel 数据。"""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return ObservationTable()
            rows_iter = ws.iter_rows(values_only=True)
            all_rows = list(rows_iter)
            if not all_rows:
                return ObservationTable()

            columns: tuple[str, ...] = tuple(
                str(c) if c is not None else f"col_{i}" for i, c in enumerate(all_rows[0])
            )
            data_rows: list[dict[str, Any]] = []
            for row in all_rows[1:]:
                if all(cell is None for cell in row):
                    continue
                record: dict[str, Any] = {}
                for col_name, cell in zip(columns, row, strict=False):
                    record[col_name] = cell
                data_rows.append(record)

            return ObservationTable(
                columns=columns,
                rows=tuple(data_rows),
                source_locations=({"object_key": object_key},),
            )
        finally:
            wb.close()
