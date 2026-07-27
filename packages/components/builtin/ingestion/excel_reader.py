"""Excel 读取组件。

使用 openpyxl 读取 .xlsx 文件，输出 ObservationTable。

参数：
- path: xlsx 文件路径（必填）。
- sheet_name: 工作表名称（可选，默认第一个工作表）。
- header_row: 表头所在行号，1-based（可选，默认 1）。
- data_start_row: 数据起始行号，1-based（可选，默认 header_row+1）。
"""

import asyncio
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from packages.components.builtin.types import ObservationTable
from packages.components.sdk import ComponentContext, ComponentResult


def _read_excel_sync(
    path_str: str,
    sheet_name: str | None,
    header_row: int,
    data_start_row: int,
) -> tuple[tuple[str, ...], list[dict[str, Any]], list[dict[str, Any]], str | None]:
    """同步读取 Excel 文件（在线程池中执行，F-21）。

    Returns:
        (columns, data_rows, source_locs, sheet_title_or_none)
        sheet_title 为 None 表示无可用工作表。
    """
    wb = load_workbook(Path(path_str), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            return (), [], [], None

        rows_iter = ws.iter_rows(
            min_row=header_row,
            values_only=True,
        )
        all_rows = list(rows_iter)
        if not all_rows:
            return (), [], [], ""

        # 表头
        header = [
            str(c).strip() if c is not None else f"col_{i}"
            for i, c in enumerate(all_rows[0])
        ]
        columns: tuple[str, ...] = tuple(header)

        data_rows: list[dict[str, Any]] = []
        source_locs: list[dict[str, Any]] = []
        offset = data_start_row - (header_row + 1)
        for idx, row in enumerate(all_rows[1 + offset :], start=data_start_row):
            if all(cell is None for cell in row):
                continue
            record: dict[str, Any] = {}
            for col_name, cell in zip(columns, row):
                record[col_name] = cell
            data_rows.append(record)
            source_locs.append(
                {
                    "file": Path(path_str).name,
                    "sheet": ws.title,
                    "row": idx,
                }
                )
        return columns, data_rows, source_locs, ws.title
    finally:
        wb.close()


class ExcelReader:
    """Excel (.xlsx) 文件读取组件。"""

    async def execute(
        self,
        context: ComponentContext,
        params: dict[str, Any],
    ) -> ComponentResult:
        """读取 xlsx 文件并输出 ObservationTable。"""
        path_str: str = params["path"]
        sheet_name: str | None = params.get("sheet_name")
        header_row: int = int(params.get("header_row", 1))
        data_start_row: int = int(
            params.get("data_start_row", header_row + 1)
        )

        # F-21: 同步文件 I/O 放 asyncio.to_thread() 避免阻塞事件循环
        columns, data_rows, source_locs, sheet_title = (
            await asyncio.to_thread(
                _read_excel_sync,
                path_str,
                sheet_name,
                header_row,
                data_start_row,
            )
        )

        if sheet_title is None:
            return ComponentResult(
                outputs={},
                summary="工作簿中无可用工作表",
                metadata={"row_count": 0},
                diagnostics={"warnings": ["no_worksheet"]},
            )

        if sheet_title == "":
            return ComponentResult(
                outputs={},
                summary="空工作表",
                metadata={"row_count": 0},
            )

        table = ObservationTable(
            columns=columns,
            rows=tuple(data_rows),
            source_locations=tuple(source_locs),
        )
        return ComponentResult(
            outputs={"observations": table},
            summary=f"从 {Path(path_str).name} 读取 {table.row_count()} 行",
            metadata={
                "row_count": table.row_count(),
                "column_count": table.column_count(),
                "sheet": sheet_title or sheet_name or "active",
            },
        )
